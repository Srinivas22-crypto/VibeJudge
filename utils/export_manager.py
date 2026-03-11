import json
import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_DIR = "data/exports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def export_json(data: dict, podcast_id: str) -> str:
    """Export analysis result as JSON."""
    path = os.path.join(OUTPUT_DIR, f"{podcast_id}_export.json")
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"✅ JSON exported: {path}")
    return path

def export_csv(data: dict, podcast_id: str) -> str:
    """Export segment-level data as CSV."""
    segments = data.get("segments", [])
    path = os.path.join(OUTPUT_DIR, f"{podcast_id}_transcript.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["start", "end", "text"])
        writer.writeheader()
        for seg in segments:
            writer.writerow({
                "start": round(seg.get("start", 0), 2),
                "end": round(seg.get("end", 0), 2),
                "text": seg.get("text", "").strip()
            })
    print(f"✅ CSV exported: {path}")
    return path

def export_xlsx(data: dict, podcast_id: str) -> str:
    """Export full analysis to Excel with multiple sheets."""
    path = os.path.join(OUTPUT_DIR, f"{podcast_id}_report.xlsx")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    ws1["A1"] = "VibeJudge Analysis Report"
    ws1["A1"].font = Font(bold=True, size=16)
    ws1.merge_cells("A1:B1")

    rows = [
        ("Podcast ID", podcast_id),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Overall Sentiment",
         data.get("sentiment", {}).get("overall_sentiment", "N/A")),
        ("Bias Level",
         data.get("bias", {}).get("bias_level", "N/A")),
        ("Bias Score",
         data.get("bias", {}).get("overall_bias_score", "N/A")),
        ("Word Count",
         len(data.get("transcript", "").split())),
    ]
    for r, (k, v) in enumerate(rows, start=3):
        ws1.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=r, column=2, value=str(v))

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 30

    # ── Sheet 2: Transcript ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Transcript")
    headers = ["Start (s)", "End (s)", "Text"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, seg in enumerate(data.get("segments", []), start=2):
        ws2.cell(row=row_idx, column=1, value=round(seg.get("start", 0), 2))
        ws2.cell(row=row_idx, column=2, value=round(seg.get("end", 0), 2))
        ws2.cell(row=row_idx, column=3, value=seg.get("text", "").strip())

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 80

    # ── Sheet 3: Bias Flags ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Bias Flags")
    bias_headers = ["Keyword", "Category", "Severity", "Sentence", "Timestamp (s)"]
    for col, h in enumerate(bias_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, flag in enumerate(
        data.get("bias", {}).get("bias_flags", []), start=2
    ):
        ws3.cell(row=row_idx, column=1, value=flag.get("keyword", ""))
        ws3.cell(row=row_idx, column=2, value=flag.get("category", ""))
        ws3.cell(row=row_idx, column=3, value=flag.get("severity", ""))
        ws3.cell(row=row_idx, column=4, value=flag.get("sentence", "")[:200])
        ws3.cell(row=row_idx, column=5, value=flag.get("timestamp", "N/A"))

    for col in ["A", "B", "C", "D", "E"]:
        ws3.column_dimensions[col].width = 20 if col != "D" else 60

    wb.save(path)
    print(f"✅ XLSX exported: {path}")
    return path
